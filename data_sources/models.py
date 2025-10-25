from django.db import models
from django.contrib.auth.models import User
from django.utils import timezone
import os


def data_file_upload_path(instance, filename):
    """
    Генерирует путь для сохранения файла данных
    Формат: data_files/YYYY/MM/DD/filename
    """
    date = instance.upload_date
    return f'data_files/{date.year}/{date.month:02d}/{date.day:02d}/{filename}'


# Алиас для обратной совместимости со старыми миграциями
csv_upload_path = data_file_upload_path


class DataUploadLog(models.Model):
    """
    Модель для журнала загрузки файлов данных (CSV, Excel)
    """
    filename = models.CharField(
        max_length=255,
        verbose_name="Наименование файла",
        help_text="Название загруженного файла"
    )
    author = models.ForeignKey(
        User,
        on_delete=models.CASCADE,
        verbose_name="Автор загрузки",
        help_text="Пользователь, загрузивший файл"
    )
    upload_date = models.DateTimeField(
        default=timezone.now,
        verbose_name="Дата загрузки",
        help_text="Дата и время загрузки файла"
    )
    file_size = models.PositiveIntegerField(
        default=0,
        verbose_name="Размер файла (байт)",
        help_text="Размер загруженного файла в байтах"
    )
    # Тип файла
    FILE_TYPE_CHOICES = [
        ('csv', 'CSV файл'),
        ('xlsx', 'Excel файл (XLSX)'),
    ]
    
    file_type = models.CharField(
        max_length=10,
        choices=FILE_TYPE_CHOICES,
        default='csv',
        verbose_name="Тип файла",
        help_text="Тип загруженного файла"
    )
    
    rows_count = models.PositiveIntegerField(
        null=True,
        blank=True,
        verbose_name="Количество строк",
        help_text="Общее количество строк в файле"
    )
    columns_count = models.PositiveIntegerField(
        null=True,
        blank=True,
        verbose_name="Количество столбцов",
        help_text="Количество столбцов в файле"
    )
    status = models.CharField(
        max_length=20,
        choices=[
            ('uploading', 'Загружается'),
            ('processing', 'Обрабатывается'),
            ('completed', 'Завершено'),
            ('error', 'Ошибка'),
        ],
        default='uploading',
        verbose_name="Статус обработки",
        help_text="Текущий статус обработки файла"
    )
    error_message = models.TextField(
        blank=True,
        null=True,
        verbose_name="Сообщение об ошибке",
        help_text="Описание ошибки, если она возникла"
    )
    has_headers = models.BooleanField(
        default=False,
        verbose_name="Содержит заголовки столбцов",
        help_text="Отметьте, если первая строка содержит названия столбцов"
    )
    
    # Для Excel файлов - название листа
    sheet_name = models.CharField(
        max_length=255,
        blank=True,
        null=True,
        verbose_name="Название листа",
        help_text="Название листа Excel файла (если применимо)"
    )
    
    # Разделители столбцов
    DELIMITER_CHOICES = [
        (',', 'Запятая (,)'),
        (';', 'Точка с запятой (;)'),
        ('\t', 'Табуляция (\\t)'),
        ('|', 'Вертикальная черта (|)'),
        (' ', 'Пробел ( )'),
        (':', 'Двоеточие (:)'),
    ]
    
    delimiter = models.CharField(
        max_length=1,
        choices=DELIMITER_CHOICES,
        default=',',
        verbose_name="Разделитель столбцов",
        help_text="Символ, используемый для разделения столбцов в CSV файле"
    )
    original_file = models.FileField(
        upload_to=data_file_upload_path,
        blank=True,
        null=True,
        verbose_name="Исходный файл",
        help_text="Оригинальный файл, загруженный пользователем"
    )

    class Meta:
        verbose_name = "Журнал загрузки данных"
        verbose_name_plural = "Журнал загрузок данных"
        ordering = ['-upload_date']

    def __str__(self):
        return f"{self.filename} - {self.author.username} ({self.upload_date.strftime('%d.%m.%Y %H:%M')})"

    def save(self, *args, **kwargs):
        """
        Автоматическое заполнение полей при сохранении
        """
        # Если это новая запись и есть файл
        if not self.pk and self.original_file:
            # Автоматически устанавливаем имя файла из загруженного файла
            if not self.filename:
                self.filename = self.original_file.name
            
            # Автоматически определяем тип файла по расширению
            if not self.file_type:
                self._auto_detect_file_type()
            
            # Автоматически устанавливаем размер файла
            try:
                self.file_size = self.original_file.size
            except (OSError, ValueError):
                self.file_size = 0
            
            # Автоматически определяем количество строк и столбцов
            try:
                self._auto_detect_file_info()
            except Exception:
                # Если не удалось определить, оставляем поля пустыми
                pass
        
        # Если автор не заполнен, пытаемся получить текущего пользователя
        if not self.author_id:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            # Пытаемся получить текущего пользователя из контекста
            # Это работает только если save() вызывается из админ-панели или с передачей user
            if hasattr(self, '_current_user') and self._current_user:
                self.author = self._current_user
            else:
                # Если нет текущего пользователя, берем первого суперпользователя
                superuser = User.objects.filter(is_superuser=True).first()
                if superuser:
                    self.author = superuser
        
        super().save(*args, **kwargs)
    
    def _auto_detect_file_type(self):
        """
        Автоматически определяет тип файла по расширению
        """
        if not self.original_file:
            return
        
        filename = self.original_file.name.lower()
        if filename.endswith('.xlsx'):
            self.file_type = 'xlsx'
        elif filename.endswith('.csv'):
            self.file_type = 'csv'
        else:
            # По умолчанию считаем CSV
            self.file_type = 'csv'
    
    def _auto_detect_file_info(self):
        """
        Автоматически определяет количество строк и столбцов в файле
        """
        if not self.original_file:
            return
        
        if self.file_type == 'xlsx':
            self._auto_detect_excel_info()
        else:
            self._auto_detect_csv_info()
    
    def _auto_detect_csv_info(self):
        """
        Автоматически определяет количество строк и столбцов в CSV файле
        """
        import csv
        import io
        
        try:
            # Читаем файл
            with open(self.original_file.path, 'r', encoding='utf-8') as file:
                try:
                    content = file.read()
                except UnicodeDecodeError:
                    # Пробуем другие кодировки
                    file.seek(0)
                    try:
                        content = file.read(encoding='cp1251')
                    except UnicodeDecodeError:
                        file.seek(0)
                        content = file.read(encoding='latin-1')
            
            # Сначала автоматически определяем разделитель
            self._auto_detect_delimiter_from_content(content)
            
            # Создаем StringIO объект для csv.reader
            csv_content = io.StringIO(content)
            
            # Читаем CSV с определенным разделителем
            csv_reader = csv.reader(csv_content, delimiter=self.delimiter)
            
            rows = list(csv_reader)
            
            if rows:
                # Определяем количество строк и столбцов
                self.rows_count = len(rows)
                self.columns_count = len(rows[0]) if rows[0] else 0
                
        except Exception:
            # Если произошла ошибка, оставляем поля пустыми
            pass
    
    def _auto_detect_excel_info(self):
        """
        Автоматически определяет количество строк и столбцов в Excel файле
        """
        try:
            import openpyxl
            
            # Загружаем Excel файл
            workbook = openpyxl.load_workbook(self.original_file.path, read_only=True)
            
            # Берем первый лист
            sheet = workbook.active
            self.sheet_name = sheet.title
            
            # Определяем количество строк и столбцов
            self.rows_count = sheet.max_row
            self.columns_count = sheet.max_column
            
            workbook.close()
            
        except Exception:
            # Если произошла ошибка, оставляем поля пустыми
            pass
    
    def _auto_detect_delimiter_from_content(self, content):
        """
        Автоматически определяет разделитель из содержимого файла
        """
        # Подсчитываем количество каждого возможного разделителя
        delimiter_counts = {}
        possible_delimiters = [',', ';', '\t', '|', ':', ' ']
        
        for delimiter in possible_delimiters:
            delimiter_counts[delimiter] = content.count(delimiter)
        
        # Возвращаем разделитель с максимальным количеством
        best_delimiter = max(delimiter_counts, key=delimiter_counts.get)
        
        # Если нет разделителей, оставляем текущий
        if delimiter_counts[best_delimiter] > 0:
            self.delimiter = best_delimiter
    
    def auto_detect_delimiter(self):
        """
        Автоматически определяет разделитель в CSV файле
        """
        if not self.original_file:
            return self.delimiter
        
        from .csv_processor import CSVProcessor
        processor = CSVProcessor(self)
        detected_delimiter = processor.auto_detect_and_set_delimiter()
        return detected_delimiter
    
    def process_data(self):
        """
        Обрабатывает файл данных и сохраняет данные в базу
        """
        if self.file_type == 'xlsx':
            from .excel_processor import ExcelProcessor
            processor = ExcelProcessor(self)
            return processor.process_excel_file()
        else:
            from .csv_processor import CSVProcessor
            processor = CSVProcessor(self)
            return processor.process_csv_file()


class DataRecord(models.Model):
    """
    Модель для хранения данных из файлов (CSV, Excel)
    """
    upload_log = models.ForeignKey(
        DataUploadLog,
        on_delete=models.CASCADE,
        related_name='data_records',
        verbose_name="Журнал загрузки",
        help_text="Ссылка на запись в журнале загрузки"
    )
    row_number = models.PositiveIntegerField(
        verbose_name="Номер строки",
        help_text="Номер строки в файле (начиная с 1)"
    )
    column_number = models.PositiveIntegerField(
        verbose_name="Номер столбца",
        help_text="Номер столбца в файле (начиная с 1)"
    )
    data_column = models.ForeignKey(
        'DataColumn',
        on_delete=models.CASCADE,
        related_name='data_records',
        null=True,
        blank=True,
        verbose_name="Столбец данных",
        help_text="Ссылка на описание столбца файла"
    )
    cell_value = models.TextField(
        verbose_name="Значение ячейки",
        help_text="Содержимое ячейки файла"
    )
    created_at = models.DateTimeField(
        auto_now_add=True,
        verbose_name="Дата создания записи",
        help_text="Дата и время создания записи в базе данных"
    )

    class Meta:
        verbose_name = "Запись данных"
        verbose_name_plural = "Записи данных"
        ordering = ['upload_log', 'row_number', 'column_number']
        indexes = [
            models.Index(fields=['upload_log', 'row_number']),
            models.Index(fields=['upload_log', 'column_number']),
            models.Index(fields=['data_column']),
        ]

    def __str__(self):
        column_name = self.data_column.column_name if self.data_column else f"Столбец {self.column_number}"
        return f"{self.upload_log.filename} - Строка {self.row_number}, {column_name}: {self.cell_value[:50]}"


class DataColumn(models.Model):
    """
    Модель для хранения информации о столбцах файла данных
    """
    upload_log = models.ForeignKey(
        DataUploadLog,
        on_delete=models.CASCADE,
        related_name='data_columns',
        verbose_name="Журнал загрузки",
        help_text="Ссылка на запись в журнале загрузки"
    )
    column_number = models.PositiveIntegerField(
        verbose_name="Номер столбца",
        help_text="Порядковый номер столбца в файле (начиная с 1)"
    )
    column_name = models.CharField(
        max_length=255,
        verbose_name="Название столбца",
        help_text="Название столбца из заголовка файла"
    )
    original_name = models.CharField(
        max_length=255,
        blank=True,
        null=True,
        verbose_name="Оригинальное название",
        help_text="Оригинальное название столбца (если было изменено)"
    )
    data_type = models.CharField(
        max_length=50,
        blank=True,
        null=True,
        verbose_name="Тип данных",
        help_text="Определенный тип данных столбца (text, number, date, etc.)"
    )
    is_required = models.BooleanField(
        default=False,
        verbose_name="Обязательный столбец",
        help_text="Является ли столбец обязательным для заполнения"
    )
    description = models.TextField(
        blank=True,
        null=True,
        verbose_name="Описание столбца",
        help_text="Дополнительное описание назначения столбца"
    )
    created_at = models.DateTimeField(
        auto_now_add=True,
        verbose_name="Дата создания",
        help_text="Дата и время создания записи о столбце"
    )

    class Meta:
        verbose_name = "Столбец данных"
        verbose_name_plural = "Столбцы данных"
        ordering = ['upload_log', 'column_number']
        unique_together = ['upload_log', 'column_number']
        indexes = [
            models.Index(fields=['upload_log', 'column_number']),
            models.Index(fields=['column_name']),
        ]

    def __str__(self):
        return f"{self.upload_log.filename} - Столбец {self.column_number}: {self.column_name}"
