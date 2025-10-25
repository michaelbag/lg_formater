import openpyxl
from typing import List, Dict, Any, Optional
from django.core.exceptions import ValidationError
from .models import DataUploadLog, DataRecord, DataColumn


class ExcelProcessor:
    """
    Класс для обработки Excel файлов (XLSX)
    """
    
    def __init__(self, upload_log: DataUploadLog):
        self.upload_log = upload_log
        self.sheet_name = upload_log.sheet_name

    def process_excel_file(self) -> Dict[str, Any]:
        """
        Обрабатывает Excel файл и сохраняет данные в базу
        """
        if not self.upload_log.original_file:
            raise ValidationError("Файл не найден")
        
        try:
            # Загружаем Excel файл
            workbook = openpyxl.load_workbook(self.upload_log.original_file.path, read_only=True)
            
            # Выбираем лист
            if self.sheet_name:
                try:
                    sheet = workbook[self.sheet_name]
                except KeyError:
                    # Если указанный лист не найден, берем первый
                    sheet = workbook.active
                    self.upload_log.sheet_name = sheet.title
            else:
                sheet = workbook.active
                self.upload_log.sheet_name = sheet.title
            
            # Получаем все данные из листа
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows:
                raise ValidationError("Excel файл пуст")
            
            # Определяем количество строк и столбцов
            rows_count = len(rows)
            columns_count = len(rows[0]) if rows[0] else 0
            
            # Проверяем, что все строки имеют одинаковое количество столбцов
            for i, row in enumerate(rows):
                if len(row) != columns_count:
                    raise ValidationError(
                        f"Строка {i+1} содержит {len(row)} столбцов, "
                        f"ожидается {columns_count}"
                    )
            
            # Обновляем информацию о файле
            self.upload_log.rows_count = rows_count
            self.upload_log.columns_count = columns_count
            self.upload_log.status = 'processing'
            self.upload_log.save()
            
            # Обрабатываем заголовки
            headers = []
            if self.upload_log.has_headers and rows:
                headers = [str(cell) if cell is not None else f"Столбец_{i+1}" 
                          for i, cell in enumerate(rows[0])]
            else:
                headers = [f"Столбец_{i+1}" for i in range(columns_count)]
            
            # Создаем записи о столбцах
            self._create_column_records(headers)
            
            # Обрабатываем данные
            self._process_data_rows(rows)
            
            # Обновляем статус
            self.upload_log.status = 'completed'
            self.upload_log.save()
            
            workbook.close()
            
            return {
                'success': True,
                'rows_processed': rows_count,
                'columns_processed': columns_count,
                'message': f"Успешно обработано {rows_count} строк и {columns_count} столбцов"
            }
            
        except Exception as e:
            # Обновляем статус на ошибку
            self.upload_log.status = 'error'
            self.upload_log.error_message = str(e)
            self.upload_log.save()
            
            return {
                'success': False,
                'error': str(e)
            }
    
    def _create_column_records(self, headers: List[str]):
        """
        Создает записи о столбцах
        """
        # Удаляем существующие записи о столбцах
        DataColumn.objects.filter(upload_log=self.upload_log).delete()
        
        # Создаем новые записи
        for i, header in enumerate(headers):
            DataColumn.objects.create(
                upload_log=self.upload_log,
                column_number=i + 1,
                column_name=header,
                original_name=header
            )
    
    def _process_data_rows(self, rows: List[tuple]):
        """
        Обрабатывает строки данных и сохраняет их в базу
        """
        # Удаляем существующие записи данных
        DataRecord.objects.filter(upload_log=self.upload_log).delete()
        
        # Определяем начальную строку (пропускаем заголовки если есть)
        start_row = 2 if self.upload_log.has_headers else 1
        
        # Получаем информацию о столбцах
        columns = {col.column_number: col for col in 
                  DataColumn.objects.filter(upload_log=self.upload_log)}
        
        # Обрабатываем каждую строку
        for row_index, row in enumerate(rows[start_row-1:], start_row):
            for col_index, cell_value in enumerate(row):
                # Преобразуем значение в строку
                str_value = str(cell_value) if cell_value is not None else ""
                
                # Создаем запись данных
                DataRecord.objects.create(
                    upload_log=self.upload_log,
                    row_number=row_index,
                    column_number=col_index + 1,
                    data_column=columns.get(col_index + 1),
                    cell_value=str_value
                )
    
    def get_available_sheets(self) -> List[str]:
        """
        Возвращает список доступных листов в Excel файле
        """
        if not self.upload_log.original_file:
            return []
        
        try:
            workbook = openpyxl.load_workbook(self.upload_log.original_file.path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception:
            return []
    
    def set_sheet(self, sheet_name: str) -> bool:
        """
        Устанавливает активный лист для обработки
        """
        available_sheets = self.get_available_sheets()
        if sheet_name in available_sheets:
            self.upload_log.sheet_name = sheet_name
            self.upload_log.save()
            return True
        return False
